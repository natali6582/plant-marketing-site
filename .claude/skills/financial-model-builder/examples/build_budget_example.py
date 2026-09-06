#!/usr/bin/env python
"""12-month operating budget following the skill contract."""
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter as L

BLUE, GREEN, GRAY = Font(color="0000FF"), Font(color="008000"), Font(color="666666")
YELLOW, GRAY_FILL = PatternFill("solid", fgColor="FFFF00"), PatternFill("solid", fgColor="D9D9D9")

def build(out, months, rev0, growth, fixed, var_pct):
    wb = Workbook(); rd = wb.active; rd.title = "README"; rd.sheet_view.rightToLeft = True
    for i, t in enumerate(["תקציב תפעולי חודשי — דוגמה", "תאים לעריכה: Assumptions!C2:C5 (כחול); אופק נקבע בבנייה עם --months (אפור)", "גרסה: 0.1",
                           "תאריך בנייה: דטרמיניסטי — ראה rebuild_compare", "המודל מחשב בלבד."], 1): rd.cell(i, 1, t)
    asm = wb.create_sheet("Assumptions"); asm.sheet_view.rightToLeft = True; asm.append(["key", "label", "value", "unit", "source"])
    rows = [("rev0", "הכנסות חודש 1", rev0, "₪", "המשתמש"), ("growth_m", "צמיחה חודשית", growth, "%", "הנחה — לאישור"),
            ("fixed", "הוצאות קבועות", fixed, "₪", "המשתמש"), ("var_pct", "הוצאות משתנות % מהכנסות", var_pct, "%", "המשתמש"),
            ("horizon", "אופק (חודשים)", months, "חודשים", "CLI --months")]
    for r in rows: asm.append(r)
    for r in range(2, 2 + len(rows)):
        is_build_time = asm.cell(r, 1).value == "horizon"
        asm.cell(r, 3).font = GRAY if is_build_time else BLUE
        if is_build_time: asm.cell(r, 3).fill = GRAY_FILL
        if asm.cell(r, 4).value == "%": asm.cell(r, 3).number_format = "0.0%"
        if "לאישור" in str(asm.cell(r, 5).value) and not is_build_time: asm.cell(r, 3).fill = YELLOW
    A = {k: f"Assumptions!$C${i+2}" for i, (k, *_) in enumerate(rows)}

    c = wb.create_sheet("Calc"); c.sheet_view.rightToLeft = True
    c.append(["key", "label"] + [str(m) for m in range(1, months + 1)])
    keys = ["revenue", "exp_fixed", "exp_var", "expense", "profit"]
    labels = ["הכנסות", "הוצאות קבועות", "הוצאות משתנות", "סה\"כ הוצאות", "רווח תפעולי"]
    R = {k: i + 2 for i, k in enumerate(keys)}
    for k, lab in zip(keys, labels): c.cell(R[k], 1, k); c.cell(R[k], 2, lab); c.cell(R[k], 1).font = GREEN
    for j in range(months):
        col = j + 3; cl, prev = L(col), L(col - 1)
        c.cell(R["revenue"], col, f"={A['rev0']}" if j == 0 else f"={prev}{R['revenue']}*(1+{A['growth_m']})")
        c.cell(R["exp_fixed"], col, f"={A['fixed']}+0*{cl}{R['revenue']}")
        c.cell(R["exp_var"], col, f"={cl}{R['revenue']}*{A['var_pct']}")
        c.cell(R["expense"], col, f"=SUM({cl}{R['exp_fixed']}:{cl}{R['exp_var']})")
        c.cell(R["profit"], col, f"={cl}{R['revenue']}-{cl}{R['expense']}")
        for k in keys: c.cell(R[k], col).number_format = "#,##0;(#,##0);-"

    o = wb.create_sheet("Output"); o.sheet_view.rightToLeft = True; last = L(months + 2)
    for i, (k, lab) in enumerate(zip(["revenue", "expense", "profit"], ["סה\"כ הכנסות שנתי", "סה\"כ הוצאות שנתי", "סה\"כ רווח שנתי"]), 1):
        o.cell(i, 1, lab); o.cell(i, 2, f"=SUM(Calc!C{R[k]}:{last}{R[k]})").number_format = "#,##0"; o.cell(i, 2).font = GREEN
    o["A5"] = "ניתוח רגישות: רווח שנתי — צמיחה חודשית (שורות) × הוצאות משתנות % (עמודות)"
    vs, gs = [0.40, 0.45, 0.50, 0.55], [0.000, 0.005, 0.010, 0.015, 0.020]
    for j, v in enumerate(vs): cell = o.cell(6, 2 + j, v); cell.number_format = "0%"; cell.font = BLUE
    for i, g in enumerate(gs):
        cell = o.cell(7 + i, 1, g); cell.number_format = "0.0%"; cell.font = BLUE
        for j in range(len(vs)):
            gc, vc, N = f"$A{7+i}", f"{L(2+j)}$6", A["horizon"]
            # annual revenue = rev0 * ((1+g)^N - 1)/g  (geometric series); profit = rev*(1-v) - fixed*N
            rev = f"IF({gc}=0,{A['rev0']}*{N},{A['rev0']}*((1+{gc})^{N}-1)/{gc})"
            o.cell(7 + i, 2 + j, f"={rev}*(1-{vc})-{A['fixed']}*{N}").number_format = "#,##0"
    wb.save(out)

if __name__ == "__main__":
    ap = argparse.ArgumentParser(); ap.add_argument("--out", required=True)
    for k, v in dict(months=12, rev0=500_000, growth=0.015, fixed=180_000, var_pct=0.45).items():
        ap.add_argument(f"--{k}", type=type(v), default=v)
    a = ap.parse_args(); build(a.out, a.months, a.rev0, a.growth, a.fixed, a.var_pct)
