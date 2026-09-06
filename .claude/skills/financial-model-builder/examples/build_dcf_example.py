#!/usr/bin/env python
"""5-year DCF following the skill contract. Period 0 in column C (investment), years 1..N in D.."""
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter as L

BLUE, GREEN, GRAY = Font(color="0000FF"), Font(color="008000"), Font(color="666666")
YELLOW, GRAY_FILL = PatternFill("solid", fgColor="FFFF00"), PatternFill("solid", fgColor="D9D9D9")

def build(out, N, rev0, growth, margin, tax, dep, capex, nwc, wacc, g, inv):
    wb = Workbook(); rd = wb.active; rd.title = "README"
    for i, t in enumerate(["מודל DCF — דוגמה", "תאים לעריכה: Assumptions!C2:C11 (כחול); אופק נקבע בבנייה עם --N (אפור)", "גרסה: 0.1",
                           "תאריך בנייה: דטרמיניסטי — ראה rebuild_compare", "המודל מחשב בלבד — אינו ייעוץ השקעות."], 1): rd.cell(i, 1, t)
    asm = wb.create_sheet("Assumptions"); asm.append(["key", "label", "value", "unit", "source"])
    rows = [("rev0", "הכנסות שנת בסיס", rev0, "₪", "המשתמש"), ("growth", "צמיחת הכנסות", growth, "%", "המשתמש"),
            ("margin", "מרווח EBIT", margin, "%", "המשתמש"), ("tax_corp", "מס חברות", tax, "%", "המשתמש"),
            ("dep_pct", "פחת % מהכנסות", dep, "%", "המשתמש"), ("capex_pct", "CapEx % מהכנסות", capex, "%", "המשתמש"),
            ("nwc_pct", "הון חוזר % מהכנסות", nwc, "%", "המשתמש"), ("wacc", "WACC", wacc, "%", "הנחה — לאישור"),
            ("g_terminal", "צמיחה טרמינלית", g, "%", "הנחה — לאישור"), ("investment", "השקעה ראשונית", inv, "₪", "המשתמש"),
            ("horizon", "אופק (שנים)", N, "שנים", "CLI --N")]
    for r in rows: asm.append(r)
    for r in range(2, 2 + len(rows)):
        is_build_time = asm.cell(r, 1).value == "horizon"
        asm.cell(r, 3).font = GRAY if is_build_time else BLUE
        if is_build_time: asm.cell(r, 3).fill = GRAY_FILL
        if asm.cell(r, 4).value == "%": asm.cell(r, 3).number_format = "0.0%"
        if "לאישור" in str(asm.cell(r, 5).value) and not is_build_time: asm.cell(r, 3).fill = YELLOW
    A = {k: f"Assumptions!$C${i+2}" for i, (k, *_) in enumerate(rows)}

    c = wb.create_sheet("Calc"); c.append(["key", "label"] + [str(y) for y in range(0, N + 1)])
    keys = ["revenue", "ebit", "tax", "nopat", "dep", "capex", "nwc", "dnwc", "fcf", "df", "pv"]
    labels = ["הכנסות", "EBIT", "מס", "NOPAT", "פחת", "CapEx", "הון חוזר", "שינוי בהון חוזר", "FCF", "מקדם היוון", "ערך נוכחי"]
    R = {k: i + 2 for i, k in enumerate(keys)}
    for k, lab in zip(keys, labels): c.cell(R[k], 1, k); c.cell(R[k], 2, lab); c.cell(R[k], 1).font = GREEN
    for j in range(N + 1):
        col = j + 3; cl, prev = L(col), L(col - 1)
        c.cell(R["revenue"], col, f"={A['rev0']}" if j == 0 else f"={prev}{R['revenue']}*(1+{A['growth']})")
        c.cell(R["ebit"], col, f"={cl}{R['revenue']}*{A['margin']}")
        c.cell(R["tax"], col, f"={cl}{R['ebit']}*{A['tax_corp']}")
        c.cell(R["nopat"], col, f"={cl}{R['ebit']}-{cl}{R['tax']}")
        c.cell(R["dep"], col, f"={cl}{R['revenue']}*{A['dep_pct']}")
        c.cell(R["capex"], col, f"={cl}{R['revenue']}*{A['capex_pct']}")
        c.cell(R["nwc"], col, f"={cl}{R['revenue']}*{A['nwc_pct']}")
        c.cell(R["dnwc"], col, f"=0*{cl}{R['nwc']}" if j == 0 else f"={cl}{R['nwc']}-{prev}{R['nwc']}")
        c.cell(R["fcf"], col, f"=-{A['investment']}" if j == 0 else f"={cl}{R['nopat']}+{cl}{R['dep']}-{cl}{R['capex']}-{cl}{R['dnwc']}")
        c.cell(R["df"], col, f"=1/(1+{A['wacc']})^{cl}$1")
        c.cell(R["pv"], col, f"={cl}{R['fcf']}*{cl}{R['df']}")
        for k in keys: c.cell(R[k], col).number_format = "0.000" if k == "df" else "#,##0;(#,##0);-"

    o = wb.create_sheet("Output"); last = L(N + 3); first = "C"
    o["A1"], o["B1"] = "סה\"כ ערך נוכחי של התזרימים", f"=SUM(Calc!{first}{R['pv']}:{last}{R['pv']})"
    o["A2"], o["B2"] = "ערך טרמינלי (מהוון)", f"=IF({A['wacc']}>{A['g_terminal']},Calc!{last}{R['fcf']}*(1+{A['g_terminal']})/({A['wacc']}-{A['g_terminal']})*Calc!{last}{R['df']},NA())"
    o["A3"], o["B3"] = "NPV (כולל השקעה)", "=B1+B2"
    o["A4"], o["B4"] = "IRR (ללא ערך טרמינלי)", f"=IRR(Calc!{first}{R['fcf']}:{last}{R['fcf']})"
    for r in (1, 2, 3): o.cell(r, 2).number_format = "#,##0"; o.cell(r, 2).font = GREEN
    o["B4"].number_format = "0.0%"
    o["A6"] = "ניתוח רגישות: NPV — WACC (שורות) × צמיחה טרמינלית (עמודות)"
    gs, ws = [0.00, 0.01, 0.02, 0.03], [0.08, 0.09, 0.10, 0.11, 0.12]
    for j, gg in enumerate(gs): cell = o.cell(7, 2 + j, gg); cell.number_format = "0.0%"; cell.font = BLUE
    fcf_rng, per_rng = f"Calc!${first}${R['fcf']}:${last}${R['fcf']}", f"Calc!${first}$1:${last}$1"
    for i, w in enumerate(ws):
        cell = o.cell(8 + i, 1, w); cell.number_format = "0.0%"; cell.font = BLUE
        for j in range(len(gs)):
            wc, gc = f"$A{8+i}", f"{L(2+j)}$7"
            o.cell(8 + i, 2 + j, f"=IF({wc}>{gc},SUMPRODUCT({fcf_rng}/(1+{wc})^{per_rng})+Calc!${last}${R['fcf']}*(1+{gc})/({wc}-{gc})/(1+{wc})^{A['horizon']},NA())").number_format = "#,##0"
    wb.save(out)

if __name__ == "__main__":
    ap = argparse.ArgumentParser(); ap.add_argument("--out", required=True)
    for k, v in dict(N=5, rev0=10_000_000, growth=0.08, margin=0.20, tax=0.23, dep=0.04, capex=0.05, nwc=0.10, wacc=0.10, g=0.02, inv=8_000_000).items():
        ap.add_argument(f"--{k}", type=type(v), default=v)
    a = ap.parse_args(); build(a.out, a.N, a.rev0, a.growth, a.margin, a.tax, a.dep, a.capex, a.nwc, a.wacc, a.g, a.inv)
