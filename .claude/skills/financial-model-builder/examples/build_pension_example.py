#!/usr/bin/env python
"""Minimal pension roll-forward model that follows the skill contract. Inputs only via CLI; no state."""
import argparse, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter as L

BLUE, GREEN, GRAY = Font(color="0000FF"), Font(color="008000"), Font(color="666666")
YELLOW = PatternFill("solid", fgColor="FFFF00")
GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")

def build(out, years, p0, deposit, ret, fee, sal_growth):
    wb = Workbook(); rd = wb.active; rd.title = "README"
    rd.sheet_view.rightToLeft = True
    for i, t in enumerate(["מודל צבירה פנסיונית — דוגמה", "תאים לעריכה: Assumptions!C2:C6 (כחול); אופק נקבע בבנייה עם --years (אפור)",
                           "גרסה: 0.1", "תאריך בנייה: (נקבע בזמן הבנייה — ראה תא C4 של Assumptions אינו קיים; הבנייה דטרמיניסטית)",
                           "המודל מחשב בלבד — אינו ייעוץ פנסיוני."], 1): rd.cell(i, 1, t)

    asm = wb.create_sheet("Assumptions"); asm.sheet_view.rightToLeft = True
    asm.append(["key", "label", "value", "unit", "source"])
    rows = [("p0", "צבירה קיימת", p0, "₪", "המשתמש"),
            ("deposit0", "הפקדה שנתית התחלתית", deposit, "₪", "המשתמש"),
            ("sal_growth", "עליית שכר שנתית", sal_growth, "%", "הנחה — לאישור"),
            ("return_lt", "תשואה שנתית", ret, "%", "הנחה — לאישור"),
            ("fee_aum", "דמי ניהול מצבירה", fee, "%", "המשתמש"),
            ("horizon", "אופק (שנים)", years, "שנים", "CLI --years")]
    for r in rows: asm.append(r)
    for r in range(2, 2 + len(rows)):
        is_build_time = asm.cell(r, 1).value == "horizon"
        asm.cell(r, 3).font = GRAY if is_build_time else BLUE
        if is_build_time: asm.cell(r, 3).fill = GRAY_FILL
        if asm.cell(r, 4).value == "%": asm.cell(r, 3).number_format = "0.00%"
        if "לאישור" in str(asm.cell(r, 5).value) and not is_build_time: asm.cell(r, 3).fill = YELLOW
    A = {k: f"Assumptions!$C${i+2}" for i, (k, *_) in enumerate(rows)}

    c = wb.create_sheet("Calc"); c.sheet_view.rightToLeft = True
    c.append(["key", "label"] + [str(y) for y in range(1, years + 1)])
    keys = ["open", "deposit", "return", "fee", "withdraw", "tax", "close", "cost_total"]
    labels = ["יתרת פתיחה", "הפקדה", "תשואה", "דמי ניהול", "משיכה", "מס", "יתרת סגירה", "סה\"כ עלויות"]
    row = {k: i + 2 for i, k in enumerate(keys)}
    for k, lab in zip(keys, labels): c.cell(row[k], 1, k); c.cell(row[k], 2, lab)
    for j in range(years):
        col = j + 3; cl = L(col); prev = L(col - 1)
        c.cell(row["open"], col, f"={A['p0']}" if j == 0 else f"={prev}{row['close']}")
        c.cell(row["deposit"], col, f"={A['deposit0']}*(1+{A['sal_growth']})^({cl}$1-1)")
        c.cell(row["return"], col, f"={cl}{row['open']}*{A['return_lt']}")
        c.cell(row["fee"], col, f"={cl}{row['open']}*{A['fee_aum']}")
        c.cell(row["withdraw"], col, f"=0*{cl}{row['open']}")
        c.cell(row["tax"], col, f"=0*{cl}{row['withdraw']}")
        c.cell(row["close"], col, f"={cl}{row['open']}+{cl}{row['deposit']}+{cl}{row['return']}-{cl}{row['fee']}-{cl}{row['withdraw']}-{cl}{row['tax']}")
        c.cell(row["cost_total"], col, f"=SUM({cl}{row['fee']}:{cl}{row['tax']})")
        for k in keys: c.cell(row[k], col).number_format = "#,##0;(#,##0);-"
    for r in range(1, 2 + len(keys)): c.cell(r, 1).font = GREEN

    o = wb.create_sheet("Output"); o.sheet_view.rightToLeft = True
    last = L(years + 2)
    o["A1"], o["B1"] = "צבירה בסוף האופק", f"=Calc!{last}{row['close']}"
    o["B1"].number_format = "#,##0"; o["B1"].font = GREEN
    o["A3"] = "ניתוח רגישות: צבירה סופית — תשואה (שורות) × דמי ניהול (עמודות)"
    fees = [0.002, 0.004, 0.005, 0.006, 0.008]; rets = [0.02, 0.03, 0.04, 0.05, 0.06]
    for j, f in enumerate(fees): cell = o.cell(4, 2 + j, f); cell.number_format = "0.0%"; cell.font = BLUE
    for i, r in enumerate(rets):
        cell = o.cell(5 + i, 1, r); cell.number_format = "0.0%"; cell.font = BLUE
        for j in range(len(fees)):
            rc, fc = f"$A{5+i}", f"{L(2+j)}$4"
            n = f"({rc}-{fc})"; g = A["sal_growth"]; N = A["horizon"]
            # closed form with growing deposit: P0*(1+n)^N + D*((1+n)^N-(1+g)^N)/(n-g)
            o.cell(5 + i, 2 + j, f"=IFERROR({A['p0']}*(1+{n})^{N}+{A['deposit0']}*((1+{n})^{N}-(1+{g})^{N})/({n}-{g}),{A['p0']}*(1+{n})^{N}+{A['deposit0']}*{N}*(1+{g})^({N}-1))").number_format = "#,##0"
    wb.save(out)

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True); ap.add_argument("--years", type=int, default=10)
    ap.add_argument("--p0", type=float, default=100000); ap.add_argument("--deposit", type=float, default=24000)
    ap.add_argument("--ret", type=float, default=0.04); ap.add_argument("--fee", type=float, default=0.005)
    ap.add_argument("--sal_growth", type=float, default=0.02)
    a = ap.parse_args(); build(a.out, a.years, a.p0, a.deposit, a.ret, a.fee, a.sal_growth)
