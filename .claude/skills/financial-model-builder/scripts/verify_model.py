#!/usr/bin/env python
"""verify_model.py — evidence report for a financial-model xlsx (skill-side audit tool, not a site dependency).

Usage: python verify_model.py <model.xlsx> --type pension|dcf|budget --recalc PATH_TO_recalc.py

Exit codes: 0 = RESULT PASS · 1 = RESULT FAIL · 2 = RESULT ERROR (verification could not complete)
A report (verify_report.json) is written next to the model when its parent directory is writable.

Model contract (see references/checks.md):
  sheets: README, Assumptions, Calc, Output (exact names)
  Assumptions: col A = key, col B = label, col C = value, col D = unit, col E = source
  Calc:        col A = key, col B = label, cols C.. = periods (row 1 = period index header)
  Output:      headline results as formulas in col B; a sensitivity block titled "רגישות"/"Sensitivity"
"""
import json, re, subprocess, sys, argparse, shutil, math
from pathlib import Path

REQ_SHEETS = ["README", "Assumptions", "Calc", "Output"]
ALLOWED_LITERALS = {"0", "1", "12", "100", "-1", "0.5"}
BAD_FUNCS = ["XLOOKUP", "XMATCH", "FILTER(", "SORT(", "UNIQUE(", "SEQUENCE(", "LET(", "LAMBDA("]
TOTAL_WORDS = ["סה\"כ", "סה״כ", "total", "סכום"]
# B. plausibility ranges: (lo, hi, severity). Substring match on the assumption key.
RANGES_COMMON = {
    "return":       (-0.50, 0.50, "FAIL"),
    "return_lt":    (0.00, 0.12, "WARN"),
    "fee_deposit":  (0.00, 0.06, "WARN"),
    "inflation":    (-0.02, 0.10, "WARN"),
    "wacc":         (0.00, 0.25, "WARN"),
    "discount":     (0.00, 0.25, "WARN"),
    "g_terminal":   (-0.01, 0.04, "WARN"),
    "retire_age":   (60, 75, "WARN"),
    "horizon":      (1, 70, "WARN"),
    "contrib_rate": (0.00, 0.30, "WARN"),
    "tax_cg":       (0.00, 0.30, "WARN"),
    "tax_corp":     (0.00, 0.35, "WARN"),
}
# fee_aum cap depends on the product: pension fund 0.5%; gemel/hishtalmut 1.05%
RANGES_BY_TYPE = {
    "pension": {"fee_aum": (0.00, 0.005, "WARN")},
    "dcf": {}, "budget": {},
}

def strip_strings(f):
    return re.sub(r'"[^"]*"', '', re.sub(r"'[^']*'!", '', f))

def literals_in(formula):
    body = strip_strings(formula[1:])
    # SUBTOTAL's sum selector is syntax, not a financial assumption.
    body = re.sub(r"\bSUBTOTAL\(\s*(?:9|109)\s*,", "SUBTOTAL(", body, flags=re.I)
    body = re.sub(r"\$?[A-Z]{1,3}\$?\d+", "", body)
    body = re.sub(r"[A-Z_]+\d*\(", "(", body)
    return [n for n in re.findall(r"(?<![A-Za-z_])-?\d+\.?\d*", body) if n not in ALLOWED_LITERALS]

def to_r1c1(formula, col):
    from openpyxl.utils import column_index_from_string as cis
    def rep(m):
        cabs, cl, rabs, rw = m.groups(); c = cis(cl)
        c = f"C{c}" if cabs else f"C[{c-col}]"
        return f"{c}R{rw}" if rabs else f"{c}R[{rw}]"
    return re.sub(r"(\$?)([A-Z]{1,3})(\$?)(\d+)", rep, strip_strings(formula))

def is_formula(v): return isinstance(v, str) and v.startswith("=")
def num(v): return isinstance(v, (int, float)) and not isinstance(v, bool) and math.isfinite(v)

def subtotal_problem(cell, calc):
    """Validate one labeled total against the documented Calc/Output layout."""
    from openpyxl.utils.cell import range_boundaries

    if not is_formula(cell.value): return "requires a SUM/SUBTOTAL formula"
    match = re.fullmatch(r"=\s*(?:SUM\(\s*|SUBTOTAL\(\s*(?:9|109)\s*,\s*)(.*?)\s*\)\s*", cell.value, re.I)
    if not match: return "requires an unadjusted SUM or SUBTOTAL(9/109)"
    reference = match.group(1)
    sheet, address = reference.rsplit("!", 1) if "!" in reference else (cell.parent.title, reference)
    sheet = sheet.strip("'").replace("''", "'").upper()
    if not re.fullmatch(r"\$?[A-Z]{1,3}\$?[1-9]\d*:\$?[A-Z]{1,3}\$?[1-9]\d*", address, re.I):
        return "requires one explicit contiguous cell range"
    first_col, first_row, last_col, last_row = range_boundaries(address.upper())
    if first_col > last_col or first_row > last_row: return "range is reversed"
    if sheet == cell.parent.title.upper():
        if first_col == last_col == cell.column and first_row >= (2 if sheet == "CALC" else 1) and last_row < cell.row:
            return None
        return "range must be above the total in the same column"
    if cell.parent.title == "Output" and sheet == "CALC":
        if (first_row == last_row and 2 <= first_row <= calc.max_row
                and first_col == 3 and last_col == calc.max_column
                and all(calc.cell(1, col).value is not None for col in range(first_col, last_col + 1))):
            return None
        return "Output total must cover one complete Calc period row"
    return "unsupported subtotal sheet reference"

class Report:
    def __init__(s, path, mtype): s.rows, s.path, s.mtype, s.error = [], path, mtype, None
    def add(s, cid, name, status, detail=""): s.rows.append({"id": cid, "check": name, "status": status, "detail": str(detail)})
    def result(s):
        if s.error: return "ERROR"
        return "FAIL" if any(r["status"] == "FAIL" for r in s.rows) else "PASS"
    def exit_code(s): return {"PASS": 0, "FAIL": 1, "ERROR": 2}[s.result()]
    def text(s):
        out = ["=== financial-model-builder verify ===", f"file: {s.path}", f"type: {s.mtype}"]
        for r in s.rows: out.append(f"{r['id']:<4}{r['check']:.<34} {r['status']:<5} {r['detail']}")
        nf = sum(r["status"] == "FAIL" for r in s.rows); nw = sum(r["status"] == "WARN" for r in s.rows)
        if s.error: out.append(f"ERROR: {s.error}")
        out.append(f"RESULT: {s.result()} ({nf} fail, {nw} warn)")
        return "\n".join(out)
    def write(s):
        (Path(s.path).parent / "verify_report.json").write_text(json.dumps(
            {"file": str(s.path), "type": s.mtype, "result": s.result(), "error": s.error, "checks": s.rows}, ensure_ascii=False, indent=1), encoding="utf-8")

class ModelCheckFailure(ValueError): pass

def run_checks(R, path, mtype, recalc):
    from openpyxl import load_workbook
    from openpyxl.formula import Tokenizer
    from openpyxl.utils import get_column_letter as L
    from openpyxl.utils.cell import range_boundaries

    # A1 recalc (LibreOffice). Hard requirement: without it, computed values are unknown.
    recalc_path = Path(recalc)
    if not recalc_path.is_file(): raise RuntimeError(f"recalc script not found or not a file: {recalc} — pass --recalc PATH (needs LibreOffice)")
    if not shutil.which("soffice") and not shutil.which("libreoffice"): raise RuntimeError("LibreOffice (soffice) not found on PATH — required for recalculation")
    try:
        rc = subprocess.run([sys.executable, "-B", str(recalc_path), str(path)], capture_output=True, text=True)
    except OSError as e:
        raise RuntimeError(f"could not start recalc: {e}") from e
    if rc.returncode != 0:
        detail = (rc.stderr or rc.stdout or "no output").strip()
        raise RuntimeError(f"recalc exited with code {rc.returncode}: {detail[:200]}")
    try:
        rj = json.loads(rc.stdout)
    except (json.JSONDecodeError, TypeError) as e:
        detail = (rc.stdout + rc.stderr).strip() or "no output"
        raise RuntimeError("recalc produced invalid JSON: " + detail[:200]) from e
    if not isinstance(rj, dict): raise RuntimeError("recalc JSON must be an object")
    if "error" in rj: raise RuntimeError("recalc error: " + str(rj["error"])[:200])
    for field in ("total_formulas", "total_errors"):
        if field not in rj: raise RuntimeError(f"recalc JSON missing field: {field}")
        if not isinstance(rj[field], int) or isinstance(rj[field], bool): raise RuntimeError(f"recalc JSON field {field} must be an integer")
        if rj[field] < 0: raise RuntimeError(f"recalc JSON field {field} must be non-negative")
    if rj["total_formulas"] == 0: raise RuntimeError("recalc reported zero formulas; recalculation is not proven")
    R.add("A1", "recalc", "PASS" if rj["total_errors"] == 0 else "FAIL", f"formulas={rj['total_formulas']}, errors={rj['total_errors']}")

    wbf = load_workbook(path); wbv = load_workbook(path, data_only=True)

    # A2 sheets — stop here if missing, but the report is still written by main()
    missing = [s for s in REQ_SHEETS if s not in wbf.sheetnames]
    R.add("A2", "required-sheets", "PASS" if not missing else "FAIL", ",".join(missing))
    if missing: return

    calc, out, asm, rd = wbf["Calc"], wbf["Output"], wbf["Assumptions"], wbf["README"]
    calcv, asmv, outv = wbv["Calc"], wbv["Assumptions"], wbv["Output"]

    # A3 / A8 / A11
    hard, bad, div = [], [], []
    for ws in (calc, out):
        for row in ws.iter_rows():
            for c in row:
                if is_formula(c.value):
                    f = c.value.upper(); ref = f"{ws.title}!{c.coordinate}"
                    if literals_in(f): hard.append(f"{ref} {c.value}")
                    if any(b in f for b in BAD_FUNCS): bad.append(ref)
                    fd = re.sub(r"/\(1\+[^)]*\)", "", strip_strings(f))
                    if "/" in fd and "IFERROR" not in f and not re.search(r"IF\([^,]*[<>=]", f): div.append(ref)
    R.add("A3", "hardcoded-in-formula", "PASS" if not hard else "FAIL", "; ".join(hard[:3]))
    R.add("A8", "unsupported-functions", "PASS" if not bad else "FAIL", ",".join(bad[:5]))
    R.add("A11", "division-guard", "PASS" if not div else "WARN", ",".join(div[:5]))

    # A4 row consistency (from 2nd period on)
    incons = []
    for row in calc.iter_rows(min_row=2):
        fs = [(c.column, c.value) for c in row[2:] if is_formula(c.value)][1:]
        if len(fs) >= 2 and len({to_r1c1(v, col) for col, v in fs}) > 1: incons.append(f"Calc!{row[0].row}")
    R.add("A4", "row-consistency", "PASS" if not incons else "FAIL", ",".join(incons[:5]))

    # A13 keyed Calc rows: period cells are formulas or empty
    hv = [f"Calc!{c.coordinate}" for row in calc.iter_rows(min_row=2) if row[0].value for c in row[2:] if c.value is not None and not is_formula(c.value)]
    R.add("A13", "period-cells-are-formulas", "PASS" if not hv else "FAIL", ",".join(hv[:5]))

    # A5 / A6 / B
    all_f = " ".join(c.value for ws in (calc, out) for row in ws.iter_rows() for c in row if is_formula(c.value)).upper()
    ranges = {**RANGES_COMMON, **RANGES_BY_TYPE[mtype]}
    orphan, nosrc, rng, avals = [], [], [], {}
    for r in range(2, asm.max_row + 1):
        key, val, src = asm.cell(r, 1).value, asm.cell(r, 3).value, asm.cell(r, 5).value
        if key is None or val is None: continue
        avals[str(key)] = val
        if f"ASSUMPTIONS!$C${r}" not in all_f and f"ASSUMPTIONS!C{r}" not in all_f: orphan.append(str(key))
        if not src: nosrc.append(str(key))
        for k, (lo, hi, sev) in ranges.items():
            if k in str(key) and num(val) and not (lo <= val <= hi): rng.append((sev, f"{key}={val} (range {lo}..{hi})"))
    R.add("A5", "orphan-assumption", "PASS" if not orphan else "WARN", ",".join(orphan))
    R.add("A6", "assumption-source", "PASS" if not nosrc else "WARN", ",".join(nosrc))
    R.add("B", "plausibility-ranges", "FAIL" if any(s == "FAIL" for s, _ in rng) else ("WARN" if rng else "PASS"), "; ".join(d for _, d in rng))

    # A7: Calc totals occupy C..; Output labels are in A and totals start in B.
    badtot = []
    for ws in (calc, out):
        for row in ws.iter_rows():
            labels = row[:2] if ws is calc else row[:1]
            label = " ".join(str(c.value) for c in labels if c.value).lower()
            if not any(word in label for word in TOTAL_WORDS): continue
            cells = list(row[2:]) if ws is calc else [ws.cell(row[0].row, 2)] + [c for c in row[2:] if c.value is not None]
            for cell in cells:
                problem = subtotal_problem(cell, calc)
                if problem: badtot.append(f"{ws.title}!{cell.coordinate}: {problem}")
    R.add("A7", "subtotals-are-formulas", "PASS" if not badtot else "FAIL", ",".join(badtot[:5]))

    # A9 sensitivity: every grid cell is independently recomputed from the explicit contract for its model type.
    def sensitivity_result():
        def finite(v): return num(v) and math.isfinite(float(v))
        def parsed(v):
            if finite(v): return float(v)
            if isinstance(v, str):
                try:
                    x = float(v)
                    return x if math.isfinite(x) else None
                except ValueError: return None
            return None
        def formula_refs(formula):
            try:
                refs = []
                for token in Tokenizer(formula).items:
                    if token.type != "OPERAND" or token.subtype != "RANGE": continue
                    raw = token.value
                    if "!" in raw:
                        sheet, address = raw.rsplit("!", 1)
                        sheet = sheet.strip("'").replace("''", "'").upper()
                    else:
                        sheet, address = None, raw
                    address = address.replace("$", "").upper()
                    if ":" in address:
                        start, end = address.split(":", 1)
                        if start == end: address = start
                    refs.append((sheet, address))
                return refs
            except Exception:
                return None
        def local_ref(refs, row, col):
            target = f"{L(col)}{row}"
            return any(sheet in (None, "OUTPUT") and address == target for sheet, address in refs)
        def exact_calc_row_range(refs, row, last_col):
            if len(refs) != 1 or refs[0][0] != "CALC": return False
            try:
                min_col, min_row, max_col, max_row = range_boundaries(refs[0][1])
            except ValueError:
                return False
            return min_row == max_row == row and min_col == 3 and max_col == last_col
        def exact_calc_cell(refs, row, col):
            return len(refs) == 1 and refs[0][0] == "CALC" and refs[0][1] == f"{L(col)}{row}"
        def integer_assumption(value):
            if not finite(value) or value < 1 or abs(value - round(value)) > 1e-9:
                return None
            return int(round(value))

        titles = [(c.row, c.column) for row in out.iter_rows() for c in row
                  if isinstance(c.value, str) and re.search("רגישות|sensitivity", c.value, re.I)]
        if len(titles) != 1:
            return "FAIL", f"expected exactly one רגישות/Sensitivity title; found {len(titles)}"
        tr, tc = titles[0]

        def contiguous_header_cols(row):
            found = []
            for col in range(tc + 1, out.max_column + 1):
                if finite(outv.cell(row, col).value): found.append(col)
                else: break
            return found

        hdr = next((row for row in range(tr + 1, min(tr + 4, out.max_row) + 1)
                    if len(contiguous_header_cols(row)) >= 3), None)
        if hdr is None: return "FAIL", "no contiguous numeric column-header row under the title"
        cols = contiguous_header_cols(hdr)
        rows = []
        for row in range(hdr + 1, out.max_row + 1):
            if finite(outv.cell(row, tc).value): rows.append(row)
            else: break
        if len(rows) < 3 or len(cols) < 3: return "FAIL", f"grid {len(rows)}x{len(cols)} < 3x3"

        row_axis = [float(outv.cell(row, tc).value) for row in rows]
        col_axis = [float(outv.cell(hdr, col).value) for col in cols]
        if len(set(row_axis)) != len(row_axis) or len(set(col_axis)) != len(col_axis):
            return "FAIL", "duplicate row or column axis values"

        cells = [(row, col) for row in rows for col in cols]
        notf = [f"{L(col)}{row}" for row, col in cells if not is_formula(out.cell(row, col).value)]
        if notf: return "FAIL", "cells are values, not formulas: " + ",".join(notf[:5])
        refs_by_cell = {(row, col): formula_refs(out.cell(row, col).value) for row, col in cells}
        parse_errors = [f"{L(col)}{row}" for row, col in cells if refs_by_cell[(row, col)] is None]
        if parse_errors: return "FAIL", "formulas could not be parsed: " + ",".join(parse_errors[:5])
        unbound = [f"{L(col)}{row}" for row, col in cells
                   if not any(sheet in ("CALC", "ASSUMPTIONS") for sheet, _ in refs_by_cell[(row, col)])]
        if unbound: return "FAIL", "formulas not bound to Calc/Assumptions: " + ",".join(unbound[:5])
        wrong_refs = [f"{L(col)}{row}" for row, col in cells
                      if not local_ref(refs_by_cell[(row, col)], row, tc)
                      or not local_ref(refs_by_cell[(row, col)], hdr, col)]
        if wrong_refs: return "FAIL", "formulas do not reference their axis cells: " + ",".join(wrong_refs[:5])
        values = [outv.cell(row, col).value for row, col in cells]
        uncomputed = [f"{L(col)}{row}" for (row, col), value in zip(cells, values) if not finite(value)]
        if uncomputed: return "FAIL", "uncomputed cells: " + ",".join(uncomputed[:5])

        assumption_rows, calc_rows = {}, {}
        for row in range(2, asmv.max_row + 1):
            key = asmv.cell(row, 1).value
            if key is not None: assumption_rows.setdefault(str(key), []).append(row)
        for row in range(2, calcv.max_row + 1):
            key = calcv.cell(row, 1).value
            if key is not None: calc_rows.setdefault(str(key), []).append(row)
        if mtype == "budget":
            required = ["growth_m", "var_pct", "rev0", "fixed", "horizon"]
            ambiguous = [key for key in required if len(assumption_rows.get(key, [])) != 1]
            if ambiguous: return "FAIL", "budget sensitivity keys must occur once: " + ",".join(ambiguous)
            inputs = {key: asmv.cell(assumption_rows[key][0], 3).value for key in required}
            missing = [key for key in required if not finite(inputs[key])]
            horizon = integer_assumption(inputs["horizon"])
            if missing or horizon is None:
                return "FAIL", "budget sensitivity inputs missing/invalid: " + ",".join(missing or ["horizon"])
            periods = [parsed(calcv.cell(1, col).value) for col in range(3, calcv.max_column + 1)]
            if periods != [float(i) for i in range(1, horizon + 1)]:
                return "FAIL", f"budget Calc periods do not equal 1..horizon ({horizon})"
            if len(calc_rows.get("profit", [])) != 1: return "FAIL", "budget Calc key must occur once: profit"
            profit_row = calc_rows["profit"][0]
            profits = [calcv.cell(profit_row, col).value for col in range(3, calcv.max_column + 1)]
            if not all(finite(value) for value in profits): return "FAIL", "budget Calc profit has uncomputed cells"
            core_base = sum(profits)
            def expected(row_value, col_value):
                revenue = inputs["rev0"] * sum((1 + row_value) ** period for period in range(horizon))
                return revenue * (1 - col_value) - inputs["fixed"] * horizon
            row_key, col_key, target = "growth_m", "var_pct", core_base
            headline_candidates = []
            for row in range(1, tr):
                formula = out.cell(row, 2).value
                refs = formula_refs(formula) if is_formula(formula) else []
                if refs is not None and exact_calc_row_range(refs, profit_row, calcv.max_column):
                    headline_candidates.append((f"B{row}", outv.cell(row, 2).value))
        elif mtype == "dcf":
            required = ["wacc", "g_terminal", "horizon"]
            ambiguous = [key for key in required if len(assumption_rows.get(key, [])) != 1]
            if ambiguous: return "FAIL", "DCF sensitivity keys must occur once: " + ",".join(ambiguous)
            inputs = {key: asmv.cell(assumption_rows[key][0], 3).value for key in required}
            missing = [key for key in required if not finite(inputs[key])]
            horizon = integer_assumption(inputs["horizon"])
            if missing or horizon is None:
                return "FAIL", "DCF sensitivity inputs missing/invalid: " + ",".join(missing or ["horizon"])
            periods = [parsed(calcv.cell(1, col).value) for col in range(3, calcv.max_column + 1)]
            if periods != [float(i) for i in range(0, horizon + 1)]:
                return "FAIL", f"DCF Calc periods do not equal 0..horizon ({horizon})"
            if len(calc_rows.get("fcf", [])) != 1: return "FAIL", "DCF Calc key must occur once: fcf"
            fcf_row = calc_rows["fcf"][0]
            fcfs = [calcv.cell(fcf_row, col).value for col in range(3, calcv.max_column + 1)]
            if not all(finite(value) for value in fcfs): return "FAIL", "DCF Calc fcf has uncomputed cells"
            def expected(row_value, col_value):
                if 1 + row_value <= 0 or row_value <= col_value: return None
                explicit = sum(value / (1 + row_value) ** period for value, period in zip(fcfs, periods))
                terminal = fcfs[-1] * (1 + col_value) / (row_value - col_value) / (1 + row_value) ** horizon
                return explicit + terminal
            row_key, col_key = "wacc", "g_terminal"
            target = None
            headline_candidates = [(f"B{row}", outv.cell(row, 2).value) for row in range(1, tr)
                                   if isinstance(out.cell(row, 1).value, str)
                                   and re.search(r"\bNPV\b", out.cell(row, 1).value, re.I)
                                   and is_formula(out.cell(row, 2).value)]
        elif mtype == "pension":
            required = ["p0", "deposit0", "sal_growth", "return_lt", "fee_aum", "horizon"]
            ambiguous = [key for key in required if len(assumption_rows.get(key, [])) != 1]
            if ambiguous: return "FAIL", "pension sensitivity keys must occur once: " + ",".join(ambiguous)
            inputs = {key: asmv.cell(assumption_rows[key][0], 3).value for key in required}
            missing = [key for key in required if not finite(inputs[key])]
            horizon = integer_assumption(inputs["horizon"])
            if missing or horizon is None:
                return "FAIL", "pension sensitivity inputs missing/invalid: " + ",".join(missing or ["horizon"])
            periods = [parsed(calcv.cell(1, col).value) for col in range(3, calcv.max_column + 1)]
            if periods != [float(i) for i in range(1, horizon + 1)]:
                return "FAIL", f"pension Calc periods do not equal 1..horizon ({horizon})"
            if len(calc_rows.get("close", [])) != 1:
                return "FAIL", "pension Calc key must occur once: close"
            close_row = calc_rows["close"][0]
            closes = [calcv.cell(close_row, col).value for col in range(3, calcv.max_column + 1)]
            if not all(finite(value) for value in closes):
                return "FAIL", "pension Calc close has uncomputed cells"

            def expected(row_value, col_value):
                n = row_value - col_value
                if 1 + n <= 0: return None
                growth_factor = (1 + n) ** horizon
                salary_factor = (1 + inputs["sal_growth"]) ** horizon
                if abs(n - inputs["sal_growth"]) <= 1e-12:
                    deposit_value = inputs["deposit0"] * horizon * (1 + inputs["sal_growth"]) ** (horizon - 1)
                else:
                    deposit_value = inputs["deposit0"] * (growth_factor - salary_factor) / (n - inputs["sal_growth"])
                return inputs["p0"] * growth_factor + deposit_value

            row_key, col_key = "return_lt", "fee_aum"
            target = closes[-1]
            headline_formula = out.cell(1, 2).value
            headline_refs = formula_refs(headline_formula) if is_formula(headline_formula) else []
            headline_candidates = [("B1", outv.cell(1, 2).value)] if (
                is_formula(headline_formula) and exact_calc_cell(headline_refs, close_row, calcv.max_column)
            ) else []
        else:
            return "FAIL", f"unsupported model type: {mtype}"

        def safe_expected(row_value, col_value):
            try:
                value = expected(row_value, col_value)
                return value if finite(value) else None
            except (ArithmeticError, ValueError, OverflowError):
                return None

        base_expected = safe_expected(inputs[row_key], inputs[col_key])
        if mtype == "dcf": target = base_expected
        if not finite(base_expected) or abs(base_expected - target) > 1:
            return "FAIL", f"base scenario does not reconcile to core model; delta={abs(base_expected - target) if finite(base_expected) else 'invalid'}"

        mismatches, invalid, max_error = [], [], 0.0
        for row, row_value in zip(rows, row_axis):
            for col, col_value in zip(cols, col_axis):
                want = safe_expected(row_value, col_value)
                coord, actual = f"{L(col)}{row}", float(outv.cell(row, col).value)
                if not finite(want): invalid.append(coord); continue
                error = abs(actual - want); max_error = max(max_error, error)
                if error > 1: mismatches.append(f"{coord} delta={round(error, 3)}")
        if invalid: return "FAIL", "invalid sensitivity scenarios: " + ",".join(invalid[:5])
        if mismatches: return "FAIL", "grid differs from independent calculation: " + "; ".join(mismatches[:5])

        base_rows = [row for row, value in zip(rows, row_axis) if abs(value - inputs[row_key]) < 1e-9]
        base_cols = [col for col, value in zip(cols, col_axis) if abs(value - inputs[col_key]) < 1e-9]
        if len(base_rows) != 1 or len(base_cols) != 1:
            return "FAIL", f"base axes must appear exactly once: {row_key}={len(base_rows)}, {col_key}={len(base_cols)}"
        base_cell = float(outv.cell(base_rows[0], base_cols[0]).value)
        if abs(base_cell - target) > 1: return "FAIL", "base grid cell does not equal the core-model output"

        if len(headline_candidates) != 1:
            return "FAIL", f"expected one contract-bound headline output; found {[coord for coord, _ in headline_candidates]}"
        headline_coord, headline_value = headline_candidates[0]
        if not finite(headline_value) or abs(headline_value - target) > 1:
            return "FAIL", f"contract-bound headline {headline_coord} does not equal the core-model output"
        return "PASS", f"{len(rows)}x{len(cols)} grid independently recomputed; max delta={max_error:.6g}; headline={headline_coord}"

    a9_status, a9_detail = sensitivity_result()
    R.add("A9", "sensitivity-table", a9_status, a9_detail)

    # A10 / A12
    pct = [f"Assumptions!{c.coordinate}" for row in asmv.iter_rows() for c in row if "%" in (c.number_format or "") and num(c.value) and c.value > 1.5]
    R.add("A10", "percent-as-fraction", "PASS" if not pct else "WARN", ",".join(pct))
    txt = " ".join(str(c.value) for row in rd.iter_rows() for c in row if c.value).lower()
    miss = [w for w in ["תאים לעריכה", "גרסה", "תאריך"] if w not in txt]
    R.add("A12", "readme-content", "PASS" if wbf.sheetnames[0] == "README" and not miss else "FAIL", ",".join(miss))

    # C. Numeric checks. Identified workbook defects are check FAILs; unexpected verifier faults remain ERRORs.
    keys = {calcv.cell(r, 1).value: r for r in range(2, calcv.max_row + 1) if calcv.cell(r, 1).value}
    period_cols = []
    period_problem = None
    for col in range(3, calc.max_column + 1):
        if calc.cell(1, col).value is None:
            later = next((c for c in range(col + 1, calc.max_column + 1) if calc.cell(1, c).value is not None), None)
            if later is not None: period_problem = f"Calc period headers are not contiguous: blank Calc!{L(col)}1 before Calc!{L(later)}1"
            elif period_cols:
                populated = next((c for c in range(col, calc.max_column + 1)
                                  if any(calc.cell(row, c).value is not None for row in range(2, calc.max_row + 1))), None)
                if populated is not None: period_problem = f"Calc period header missing at Calc!{L(populated)}1 for populated column"
            break
        period_cols.append(col)
    if not period_cols and not period_problem: period_problem = "Calc has no active period headers from C1"
    R.add("C0", "active-periods", "FAIL" if period_problem else "PASS", period_problem or f"{len(period_cols)} contiguous period headers")
    if period_problem: return

    def rowvals(k):
        vals = []
        for col in period_cols:
            coord = f"Calc!{L(col)}{keys[k]}"
            if not is_formula(calc.cell(keys[k], col).value):
                raise ModelCheckFailure(f"{k}: missing formula at {coord}")
            value = calcv.cell(keys[k], col).value
            if not num(value):
                raise ModelCheckFailure(f"{k}: invalid cached result {coord}")
            vals.append(value)
        return vals
    def numeric_check(cid, name, need, fn, start=0, tol=0.01):
        if not all(k in keys for k in need): R.add(cid, name, "SKIP" if cid.startswith("C2") and cid not in ("C2c",) else "FAIL", "keys missing: " + ",".join(k for k in need if k not in keys)); return
        try:
            v = {k: rowvals(k) for k in need}
        except ModelCheckFailure as e:
            R.add(cid, name, "FAIL", str(e)); return
        n = len(v[need[0]]); bad = []
        for i in range(start, n):
            try:
                delta = fn({k: v[k][i] for k in need})
            except (ArithmeticError, ValueError) as e:
                R.add(cid, name, "FAIL", f"arithmetic failed in Calc column {L(period_cols[i])}: {type(e).__name__}: {e}"); return
            if not num(delta):
                R.add(cid, name, "FAIL", f"arithmetic returned a non-finite value in Calc column {L(period_cols[i])}"); return
            if abs(delta) > tol: bad.append(L(period_cols[i]))
        R.add(cid, name, "PASS" if not bad else "FAIL", ",".join(bad[:5]))
    if mtype == "pension":
        need = ["open", "deposit", "return", "fee", "withdraw", "tax", "close"]
        numeric_check("C1", "roll-forward", need, lambda v: v["open"] + v["deposit"] + v["return"] - v["fee"] - v["withdraw"] - v["tax"] - v["close"])
        if all(k in keys for k in ("open", "close")):
            link = all(calc.cell(keys["open"], c).value == f"={L(c-1)}{keys['close']}" for c in period_cols[1:])
            R.add("C1a", "open=prev-close (link)", "PASS" if link else "FAIL")
            try:
                close_values = rowvals("close")
            except ModelCheckFailure as e:
                R.add("C1b", "no-negative-balance", "FAIL", str(e))
            else:
                neg = [L(period_cols[i]) for i, v in enumerate(close_values) if v < 0]
                R.add("C1b", "no-negative-balance", "PASS" if not neg else "WARN", ",".join(neg[:5]))
    elif mtype == "dcf":
        g, w = avals.get("g_terminal"), avals.get("wacc")
        R.add("C2a", "g<wacc", "PASS" if num(g) and num(w) and g < w else "FAIL", f"g={g}, wacc={w}")
        if "fcf" not in keys:
            R.add("C2b", "period0-negative", "FAIL", "keys missing: fcf")
        else:
            try:
                fcf_values = rowvals("fcf")
            except ModelCheckFailure as e:
                R.add("C2b", "period0-negative", "FAIL", str(e))
            else:
                f0 = fcf_values[0]
                R.add("C2b", "period0-negative", "PASS" if f0 < 0 else "FAIL", f"fcf0={f0}")
        numeric_check("C2c", "fcf=nopat+dep-capex-dnwc", ["fcf", "nopat", "dep", "capex", "dnwc"], lambda v: v["fcf"] - (v["nopat"] + v["dep"] - v["capex"] - v["dnwc"]), start=1, tol=1)
        numeric_check("C2d", "assets=liab+equity", ["assets", "liabilities", "equity"], lambda v: v["assets"] - (v["liabilities"] + v["equity"]), tol=1)
        numeric_check("C2e", "cash-roll-forward", ["cash_close", "cash_open", "fcf", "financing"], lambda v: v["cash_close"] - (v["cash_open"] + v["fcf"] + v["financing"]), start=1, tol=1)
    else:
        numeric_check("C3", "profit=rev-exp", ["revenue", "expense", "profit"], lambda v: v["revenue"] - v["expense"] - v["profit"])

def main():
    ap = argparse.ArgumentParser(); ap.add_argument("model"); ap.add_argument("--type", required=True, choices=["pension", "dcf", "budget"])
    ap.add_argument("--recalc", metavar="PATH", help="required path to the xlsx skill's recalc.py (LibreOffice headless)")
    a = ap.parse_args(); path = Path(a.model); R = Report(path, a.type)
    try:
        if not path.exists(): raise RuntimeError(f"model not found: {path}")
        if not a.recalc: raise RuntimeError("--recalc PATH is required; no default recalc script is assumed")
        run_checks(R, path, a.type, a.recalc)
    except Exception as e: R.error = f"{type(e).__name__}: {e}"
    finally:
        print(R.text())
        if path.parent.exists(): R.write()
    sys.exit(R.exit_code())

if __name__ == "__main__":
    main()
