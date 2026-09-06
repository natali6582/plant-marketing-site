"""Missing/duplicate base axes must fail even when every cached grid value is correct."""
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from pension_a9_contract_tests import make_contract_fixture, output_cell, patch_output, set_cached_value
from subtotal_contract_tests import invoke

HERE = Path(__file__).resolve().parent


def make_case(mtype, scenario, root):
    source = HERE / "fixtures" / f"{mtype}_example.xlsx"
    if mtype == "pension":
        clean = root / "clean.xlsx"
        make_contract_fixture(source, clean)
        source = clean
    workbook = load_workbook(source, data_only=True)
    asm, calc, out = workbook["Assumptions"], workbook["Calc"], workbook["Output"]
    inputs = {asm.cell(row, 1).value: asm.cell(row, 3).value for row in range(2, asm.max_row + 1)}
    hdr, row_key, col_key = {
        "pension": (4, "return_lt", "fee_aum"),
        "budget": (6, "growth_m", "var_pct"),
        "dcf": (7, "wacc", "g_terminal"),
    }[mtype]
    row_axis = [out.cell(row, 1).value for row in range(hdr + 1, hdr + 6)]
    col_axis = []
    for col in range(2, out.max_column + 1):
        value = out.cell(hdr, col).value
        if not isinstance(value, (int, float)): break
        col_axis.append(value)
    if scenario == "missing-row":
        row_axis[row_axis.index(inputs[row_key])] += 0.0003
    elif scenario == "missing-column":
        col_axis[col_axis.index(inputs[col_key])] += 0.0003
    elif scenario == "duplicate-column":
        col_axis[0] = col_axis[1]

    def expected(row_value, col_value):
        if mtype == "pension":
            # Oracle uses period-by-period balances, not the verifier's closed form.
            balance = inputs["p0"]
            for period in range(int(inputs["horizon"])):
                balance = balance * (1 + row_value - col_value) + inputs["deposit0"] * (1 + inputs["sal_growth"]) ** period
            return balance
        if mtype == "budget":
            revenue, profit = inputs["rev0"], 0
            for _ in range(int(inputs["horizon"])):
                profit += revenue * (1 - col_value) - inputs["fixed"]
                revenue *= 1 + row_value
            return profit
        fcf_row = next(row for row in range(2, calc.max_row + 1) if calc.cell(row, 1).value == "fcf")
        fcfs = [calc.cell(fcf_row, col).value for col in range(3, calc.max_column + 1)]
        return sum(fcf / (1 + row_value) ** period for period, fcf in enumerate(fcfs)) + fcfs[-1] * (1 + col_value) / (row_value - col_value) / (1 + row_value) ** inputs["horizon"]

    def mutate(xml):
        for row, value in enumerate(row_axis, start=hdr + 1):
            set_cached_value(output_cell(xml, f"A{row}"), value)
        for col, value in enumerate(col_axis, start=2):
            set_cached_value(output_cell(xml, f"{chr(64 + col)}{hdr}"), value)
        for row, rv in enumerate(row_axis, start=hdr + 1):
            for col, cv in enumerate(col_axis, start=2):
                set_cached_value(output_cell(xml, f"{chr(64 + col)}{row}"), expected(rv, cv))

    model = root / "model.xlsx"
    patch_output(source, model, mutate)
    workbook.close()
    return model


def main():
    passed, total = 0, 12
    with tempfile.TemporaryDirectory(prefix="base-axes-contract-") as directory:
        for mtype in ("pension", "budget", "dcf"):
            for scenario in ("clean", "missing-row", "missing-column", "duplicate-column"):
                name = f"{mtype}-{scenario}"
                root = Path(directory) / name
                root.mkdir()
                try:
                    code, report = invoke(make_case(mtype, scenario, root), mtype)
                    a9 = next(check for check in report["checks"] if check["id"] == "A9")
                    status = "PASS" if scenario == "clean" else "FAIL"
                    detail = "independently recomputed" if scenario == "clean" else "duplicate row or column" if scenario == "duplicate-column" else "base axes must appear exactly once"
                    assert a9["status"] == status and detail in a9["detail"], a9
                    assert code == (0 if status == "PASS" else 1), (code, report)
                    print(f"PASS {name} A9={status} exit={code} deterministic=yes")
                    passed += 1
                except AssertionError as error:
                    print(f"FAIL {name}: {error}")
    print(f"SUMMARY {passed}/{total} PASS (stub; independent cached grids)")
    raise SystemExit(passed != total)


if __name__ == "__main__":
    main()
