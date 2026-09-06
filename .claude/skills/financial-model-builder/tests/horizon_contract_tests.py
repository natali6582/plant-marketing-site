import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
EXAMPLES = HERE.parent / "examples"


CASES = [
    {
        "name": "pension",
        "script": EXAMPLES / "build_pension_example.py",
        "flag": "--years",
        "horizon": 7,
        "periods": list(range(1, 8)),
        "editable": "Assumptions!C2:C6",
    },
    {
        "name": "dcf",
        "script": EXAMPLES / "build_dcf_example.py",
        "flag": "--N",
        "horizon": 4,
        "periods": list(range(0, 5)),
        "editable": "Assumptions!C2:C11",
    },
    {
        "name": "budget",
        "script": EXAMPLES / "build_budget_example.py",
        "flag": "--months",
        "horizon": 9,
        "periods": list(range(1, 10)),
        "editable": "Assumptions!C2:C5",
    },
]


def run_case(case, root):
    output = root / f"{case['name']}.xlsx"
    completed = subprocess.run(
        [sys.executable, "-B", str(case["script"]), "--out", str(output), case["flag"], str(case["horizon"])],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    if completed.returncode != 0:
        raise AssertionError(f"builder exited {completed.returncode}: {completed.stderr}")
    if not output.is_file():
        raise AssertionError("builder did not create an xlsx")

    workbook = load_workbook(output, data_only=False)
    assumptions = workbook["Assumptions"]
    horizon_rows = [r for r in range(2, assumptions.max_row + 1) if assumptions.cell(r, 1).value == "horizon"]
    if horizon_rows != [assumptions.max_row]:
        raise AssertionError(f"horizon row is not the final assumption row: {horizon_rows}")
    horizon_row = horizon_rows[0]
    horizon_cell = assumptions.cell(horizon_row, 3)
    if horizon_cell.value != case["horizon"]:
        raise AssertionError(f"horizon value is {horizon_cell.value}, expected {case['horizon']}")
    if horizon_cell.font.color is None or horizon_cell.font.color.type != "rgb" or horizon_cell.font.color.rgb[-6:] == "0000FF":
        raise AssertionError("horizon is still styled as a blue editable input")
    if horizon_cell.fill.fgColor.rgb[-6:] != "D9D9D9":
        raise AssertionError("horizon does not have the grey build-time fill")

    readme = " ".join(str(cell.value) for row in workbook["README"].iter_rows() for cell in row if cell.value)
    if case["editable"] not in readme or case["flag"] not in readme:
        raise AssertionError(f"README does not document build-time horizon/editable range: {readme}")

    calc = workbook["Calc"]
    periods = [int(calc.cell(1, col).value) for col in range(3, calc.max_column + 1)]
    if periods != case["periods"]:
        raise AssertionError(f"Calc periods {periods} do not match build-time horizon {case['horizon']}")


def main():
    with tempfile.TemporaryDirectory(prefix="horizon-contract-") as temp:
        root = Path(temp)
        for case in CASES:
            run_case(case, root)
            print(f"PASS {case['name']} horizon={case['horizon']} periods={case['periods'][0]}..{case['periods'][-1]}")
    print(f"SUMMARY {len(CASES)}/{len(CASES)} PASS")


if __name__ == "__main__":
    main()
