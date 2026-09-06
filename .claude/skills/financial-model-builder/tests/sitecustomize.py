import math
import os


mode = os.environ.get("VERIFY_TEST_SPECIAL_CACHE")
if mode:
    import openpyxl

    original_load_workbook = openpyxl.load_workbook

    def load_workbook_with_special_cache(*args, **kwargs):
        workbook = original_load_workbook(*args, **kwargs)
        if kwargs.get("data_only"):
            workbook["Calc"]["C5"].value = math.nan if mode == "nan" else math.inf
        return workbook

    openpyxl.load_workbook = load_workbook_with_special_cache
