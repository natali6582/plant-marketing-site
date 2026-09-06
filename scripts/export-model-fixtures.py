#!/usr/bin/env python
"""Export the approved pension AUM-only workbooks' cached values as site fixtures.

Manual developer tool: Python + openpyxl (also used by the model skill).
Neither Python nor the workbook is a site build/runtime dependency.
Defaults point to the versioned source pairs; no recalculation is performed.
"""
import argparse
import hashlib
import io
import json
import math
import os
from pathlib import Path
import sys
import tempfile

WORKBOOK_SHA256 = "061a3c8e95d20439bb0268247bcf003e25597de6cc13223001358d78eb3cca73"
REPORT_SHA256 = "b1f7be46eac1b29c7a298ed4773ae1d57dc6eb79ec11449184b9189049750f6c"
SOURCE_DIR = "docs/models/pension-aum-v1"
EXTRA_DIR = f"{SOURCE_DIR}/extra"
EXTRA_ARCHIVE = "pension-aum-v1-fixtures-extra.zip"
EXTRA_ARCHIVE_SHA256 = "08221ad309480636518a3a6b6490b7fa6e74f312e8ab4f01a2c61b8f89b4ce91"
EXTRA_MANIFEST_SHA256 = "5803cb7af0ddf53cdd9827dac261d8096d531d140d30ac386683eeb348aa7544"
EXTRA_IDS = ("years-30", "deposit-0", "p0-0", "salgrowth-0", "n-equals-g")
INPUT_KEYS = (("p0", "p0"), ("deposit", "deposit0"), ("salaryGrowth", "sal_growth"),
              ("ret", "return_lt"), ("feeAum", "fee_aum"), ("years", "horizon"))


def require(condition, message):
    if not condition:
        raise ValueError(message)


def checked_bytes(path, expected, name):
    try:
        data = path.read_bytes()
    except OSError as error:
        raise ValueError(f"cannot read {name}: {path}") from error
    require(hashlib.sha256(data).hexdigest() == expected, f"{name} SHA-256 does not match the approved source")
    return data


def finite_number(value, address):
    require(isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(value),
            f"missing or non-finite cached number at {address}")
    return value


def extract(workbook_bytes, report):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ValueError("openpyxl is required for this manual export; use the model-skill Python environment") from error
    require(report.get("type") == "pension" and report.get("result") == "PASS" and report.get("error") is None,
            "verification report is not a successful pension audit")
    checks = report.get("checks", [])
    expected_ids = {"A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9", "A10", "A11", "A12", "A13", "B", "C0", "C1", "C1a", "C1b"}
    require(len(checks) == 18 and {check.get("id") for check in checks} == expected_ids
            and all(check.get("status") == "PASS" for check in checks), "expected 18 distinct PASS checks")
    values = load_workbook(io.BytesIO(workbook_bytes), data_only=True, read_only=True)
    formulas = load_workbook(io.BytesIO(workbook_bytes), data_only=False, read_only=True)
    try:
        assumptions, addresses = {}, {}
        for row in values["Assumptions"].iter_rows(min_row=2):
            key = row[0].value
            if key is None: continue
            require(key not in assumptions, f"duplicate assumption key: {key}")
            address = f"Assumptions!{row[2].coordinate}"
            assumptions[key] = finite_number(row[2].value, address)
            addresses[key] = address
        require(set(assumptions) == {key for _, key in INPUT_KEYS}, "unexpected assumptions: A3 supports AUM fees only")
        base = {name: assumptions[key] for name, key in INPUT_KEYS}
        require(base["years"] >= 1 and int(base["years"]) == base["years"], "horizon must be a positive integer")
        base["years"] = int(base["years"])
        calc = values["Calc"]
        require([int(calc.cell(1, col).value) for col in range(3, calc.max_column + 1)] == list(range(1, base["years"] + 1)),
                "Calc periods disagree with the horizon")

        def cached(address):
            sheet, cell = address.split("!")
            require(formulas[sheet][cell].data_type == "f", f"expected a formula at {address}")
            return finite_number(values[sheet][cell].value, address)

        out = values["Output"]
        # This exporter is deliberately pinned to the reviewed 5x5 v1 workbook.
        returns = [(row, finite_number(out.cell(row, 1).value, f"Output!A{row}")) for row in range(5, 10)]
        fees = [(col, finite_number(out.cell(4, col).value, f"Output!{chr(64 + col)}4")) for col in range(2, 7)]
        base_rows = [row for row, value in returns if math.isclose(value, base["ret"], abs_tol=1e-12, rel_tol=0)]
        base_cols = [col for col, value in fees if math.isclose(value, base["feeAum"], abs_tol=1e-12, rel_tol=0)]
        require(len(base_rows) == len(base_cols) == 1, "base scenario must appear exactly once on both axes")
        base_row, base_col = base_rows[0], base_cols[0]
        require(returns[0][0] < base_row < returns[-1][0] and fees[0][0] < base_col < fees[-1][0], "base scenario must be inside the selected extremes")
        close_rows = [row for row in range(2, calc.max_row + 1) if calc.cell(row, 1).value == "close"]
        require(len(close_rows) == 1, "Calc must contain one close row")
        headline = cached("Output!B1")
        final_cell = calc.cell(close_rows[0], calc.max_column)
        require(formulas["Output"]["B1"].value == f"=Calc!{final_cell.coordinate}",
                "headline must link to the final keyed close")
        close = cached(f"Calc!{final_cell.coordinate}")
        base_grid = f"Output!{chr(64 + base_col)}{base_row}"
        require(abs(headline - close) <= 1 and abs(headline - cached(base_grid)) <= 1,
                "headline, final close and base grid cell do not reconcile")

        def scenario(row, col, case_id):
            inputs = dict(base)
            inputs["ret"] = out.cell(row, 1).value
            inputs["feeAum"] = out.cell(4, col).value
            cell = "Output!B1" if case_id == "base" else f"Output!{chr(64 + col)}{row}"
            return {"id": case_id, "inputs": inputs, "expectedClosingBalance": cached(cell),
                    "source": {"closingBalanceCell": cell, "returnCell": f"Output!A{row}",
                               "feeAumCell": f"Output!{chr(64 + col)}4"}}

        fixtures = [scenario(base_row, base_col, "base")]
        for row_label, row in (("low", returns[0][0]), ("base", base_row), ("high", returns[-1][0])):
            for col_label, col in (("low", fees[0][0]), ("base", base_col), ("high", fees[-1][0])):
                if (row, col) != (base_row, base_col):
                    fixtures.append(scenario(row, col, f"return-{row_label}-fee-{col_label}"))
        return {
            "schemaVersion": 1, "model": "pension-aum-v1", "currency": "ILS", "toleranceIls": 1,
            "contract": {"supportedInputs": [name for name, _ in INPUT_KEYS], "period": "year",
                         "rates": "fractions", "depositTiming": "end-of-period",
                         "returnAndAumFeeBasis": "opening-balance", "excluded": ["feeDeposit", "withdrawals", "taxes"],
                         "assumptions": "illustrative, not regulatory limits"},
            "coverage": {"variedInputs": ["ret", "feeAum"],
                         "fixedInputs": {key: base[key] for key in ("p0", "deposit", "salaryGrowth", "years")},
                         "note": "These Excel fixtures do not validate changes to the fixed inputs or excluded features."},
            "source": {"workbook": f"{SOURCE_DIR}/pension_example_recalc.xlsx", "workbookSha256": WORKBOOK_SHA256,
                       "report": f"{SOURCE_DIR}/verify_report.json", "reportSha256": REPORT_SHA256,
                       "verifiedAtPath": report["file"], "verificationResult": "PASS", "verificationChecks": 18,
                       "expectedValues": "original cached values from the supplied LibreOffice-recalculated workbook",
                       "baseInputCells": {name: addresses[key] for name, key in INPUT_KEYS},
                       "baseGridCell": base_grid},
            "fixtures": fixtures,
        }
    finally:
        values.close()
        formulas.close()


def append_extra(data, directory):
    """Append five headline caches, preserving the original nine fixtures exactly."""
    checked_bytes(directory / EXTRA_ARCHIVE, EXTRA_ARCHIVE_SHA256, "additional archive")
    manifest = json.loads(checked_bytes(directory / "manifest.json", EXTRA_MANIFEST_SHA256, "additional manifest"))
    require(manifest.get("model") == "pension-aum-v1", "unexpected additional model")
    require([case["id"] for case in manifest["scenarios"]] == list(EXTRA_IDS), "unexpected additional scenarios")
    for case in manifest["scenarios"]:
        case_id = case["id"]
        # The pinned manifest is authoritative, and paths still stay explicit.
        require(case["workbook"] == f"{case_id}/{case_id}.xlsx"
                and case["report"] == f"{case_id}/verify_report.json", "unexpected additional source paths")
        require(case["result"] == "PASS" and case["exit"] == 0 and case["checks"] == 18
                and case["warn"] == [], f"{case_id}: manifest does not record PASS/0")
        workbook = checked_bytes(directory / case["workbook"], case["workbookSha256"], f"{case_id} workbook")
        report = json.loads(checked_bytes(directory / case["report"], case["reportSha256"], f"{case_id} report"))
        # Reuse the keyed-source reader, but take only this workbook's headline.
        # Its other grid values were audited on receipt, not exported as new cases.
        sample = extract(workbook, report)
        fixture = sample["fixtures"][0]
        require(fixture["inputs"] == case["inputs"], f"{case_id}: workbook inputs differ from manifest")
        require(case["closingBalanceCell"] == "Output!B1"
                and fixture["expectedClosingBalance"] == case["expectedClosingBalance"],
                f"{case_id}: original headline cache differs from manifest")
        fixture["id"] = case_id
        fixture["source"].update({
            "workbook": f"{EXTRA_DIR}/{case['workbook']}", "workbookSha256": case["workbookSha256"],
            "report": f"{EXTRA_DIR}/{case['report']}", "reportSha256": case["reportSha256"],
            "verifiedAtPath": report["file"], "verificationResult": "PASS", "verificationChecks": 18,
            "inputCells": sample["source"]["baseInputCells"], "baseGridCell": sample["source"]["baseGridCell"],
        })
        data["fixtures"].append(fixture)
    data["coverage"] = {
        "variedInputs": [name for name, _ in INPUT_KEYS], "fixedInputs": {},
        "originalGrid": data["coverage"], "additionalFixtureIds": list(EXTRA_IDS),
        "negativeReturnsVerified": False,
        "note": "Selected variations of every supported input are covered, not every combination or numeric range. "
                "Negative returns and excluded features have no Excel fixtures. The equality case validates "
                "the Excel closed-form limit; the JS model uses an annual recurrence.",
    }
    data["additionalSource"] = {
        "archive": f"{EXTRA_DIR}/{EXTRA_ARCHIVE}", "archiveSha256": EXTRA_ARCHIVE_SHA256,
        "manifest": f"{EXTRA_DIR}/manifest.json", "manifestSha256": EXTRA_MANIFEST_SHA256,
        "expectedValues": "original Output!B1 caches from five supplied LibreOffice-recalculated workbooks",
        "nativeVerification": "Supplier reports PASS/0 and identical native runs; only run-1 artifacts supplied. "
                              "The exporter does not replay native recalculation.",
    }
    return data


def main():
    repo = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=repo / SOURCE_DIR / "pension_example_recalc.xlsx")
    parser.add_argument("--report", type=Path, default=repo / SOURCE_DIR / "verify_report.json")
    parser.add_argument("--extra-dir", type=Path, default=repo / EXTRA_DIR)
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args()
    try:
        require(args.out.suffix.lower() == ".json", "output must be a JSON file")
        require(args.out.resolve() not in (args.workbook.resolve(), args.report.resolve()), "output cannot overwrite an input")
        require(not args.out.resolve().is_relative_to(args.extra_dir.resolve()), "output cannot overwrite additional sources")
        workbook = checked_bytes(args.workbook, WORKBOOK_SHA256, "workbook")
        report = json.loads(checked_bytes(args.report, REPORT_SHA256, "report"))
        data = append_extra(extract(workbook, report), args.extra_dir)
        content = (json.dumps(data, ensure_ascii=False, indent=2, allow_nan=False) + "\n").encode("utf-8")
        args.out.parent.mkdir(parents=True, exist_ok=True)
        temporary = None
        try:
            with tempfile.NamedTemporaryFile(dir=args.out.parent, prefix=".model-fixtures-", suffix=".tmp", delete=False) as stream:
                temporary = Path(stream.name)
                stream.write(content)
            os.replace(temporary, args.out)
        finally:
            if temporary is not None:
                temporary.unlink(missing_ok=True)
        print(f"Exported {len(data['fixtures'])} pension-aum-v1 fixtures; sha256={hashlib.sha256(content).hexdigest()}")
        return 0
    except (ValueError, OSError, KeyError) as error:
        print(f"Export failed: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
